import pandas as pd
import subprocess
import sys
import os
import vcf
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Font

def load_vcf_from_string(vcf_string):
    vcf_reader = vcf.Reader(StringIO(vcf_string))
    vcf_records = []
    for record in vcf_reader:
        vcf_records.append(record)
    return vcf_records

def load_vcf_chrom_pos_bcf(vcf_file, chrom, pos):
    command = ['bcftools', 'view', '-r', f"{chrom}:{pos}", vcf_file]
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if result.returncode != 0:
        print(f"Ошибка в bcftools: {result.stderr.decode()}")
        return None
    vcf_output = result.stdout.decode()
    vcf_records = load_vcf_from_string(vcf_output)
    return vcf_records

def annotate_genotypes(excel_file, vcf_file, output_file):
    dfs = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')

    results_dict = {}

    for disease, df in dfs.items():
        column_names = list(df.columns) + ['Genotype', 'DP']

        results = []
        for _, row in df.iterrows():
            Chrom = row.iloc[0].strip()
            Pos = int(row.iloc[1])

            vcf_records = load_vcf_chrom_pos_bcf(vcf_file, Chrom, Pos)

            if vcf_records:
                record = vcf_records[0]
                genotype_info = record.genotype(record.samples[0].sample)

                genotype_str = genotype_info.gt_bases
                if genotype_str is None:
                    genotype_str = '--'
                else:
                    genotype_str = genotype_str.replace('|', '').replace('/', '')

                try:
                    dp = genotype_info['DP']
                except KeyError:
                    dp = '--'
                except Exception as e:
                    dp = '--'

                results.append(list(row) + [genotype_str, dp])
            else:
                results.append(list(row) + ['--', '--'])

        results_df = pd.DataFrame(results, columns=column_names)
        results_dict[disease] = results_df

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in results_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    workbook = load_workbook(output_file)
    for sheet_name in results_dict.keys():
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Gilroy Medium', size=12)

    workbook.save(output_file)
    print("The data is successfully saved in the output file")

excel_file = r"/mnt/tank/scratch/yunovikova/gl005/Pharmacogenomics.xlsx"
vcf_file = sys.argv[1]
output_file = os.path.join(os.path.dirname(vcf_file), os.path.splitext(os.path.splitext(vcf_file)[0])[0] + "_pharmacogenomics.xlsx")

annotate_genotypes(excel_file, vcf_file, output_file)
